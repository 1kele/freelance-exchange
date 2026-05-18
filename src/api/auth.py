from datetime import datetime

from fastapi import APIRouter, Response, Body
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from starlette.responses import FileResponse

from src.api.dependencies import DBDep, CurrentUserDep
from src.exceptions import ObjectNotFoundException, WrongPasswordException, ObjectNotFoundHTTPException, \
    WrongPasswordHTTPException, PermissionDeniedHTTPException, PermissionDeniedException, UserIsBlockedException, \
    UserIsBlockedHTTPException, UserAlreadyExistsException, UserAlreadyExistsHTTPException
from src.schemas import response
from src.schemas.order import OrderStatus
from src.schemas.user import UserAddRequest, UserLogin, UserPatch, AllRoles
from src.services.auths import AuthenticationService

router = APIRouter(prefix="/auth", tags=["Аунтефикация и Авторизация"])

@router.post("/register")
async def register(
    data: UserAddRequest,
    db: DBDep,
):
    try:
        await AuthenticationService(db).register(data)
    except UserAlreadyExistsException:
        raise UserAlreadyExistsHTTPException

    return {"status": "OK"}


@router.post("/login")
async def login(
    db: DBDep,
    response: Response,
    data: UserLogin = Body(openapi_examples={
            "1": {
                "summary": "Тест 1(Customer)",
                "value": {"email": "karate@gmail.com", "password": "RoyalOAK"},
            },
            "2": {
                "summary": "Тест 2(Freelancer)",
                "value": {"email": "kot@pes.ru", "password": "AudemarPigue"},
            },
            "3": {
                "summary": "Тест 3(Administrator)",
                "value": {"email": "avadakedavra@pes.cot", "password": "RolexDateJust"},
            },
        })
):
    try:
        access_token = await AuthenticationService(db).login(data,response)
    except WrongPasswordException:
        raise WrongPasswordHTTPException
    except ObjectNotFoundException:
        raise ObjectNotFoundHTTPException
    except UserIsBlockedException:
        raise UserIsBlockedHTTPException
    return {"access_token": access_token}


@router.get("/get_me")
async def get_me(
    db: DBDep,
    current_user: CurrentUserDep
):
    result = await AuthenticationService(db).get_me(current_user)
    return {"data": result}


@router.post("/report")
async def report(
    db: DBDep,
    date_from: datetime,
    date_to: datetime,
    current_user: CurrentUserDep,
):
    if current_user.role == AllRoles.customer:
        orders = await db.order.get_by_date_range(date_from=date_from, date_to=date_to, customer_id=current_user.id)
    elif current_user.role == AllRoles.freelancer:
        orders = await db.order.get_filter_by(date_from=date_from, date_to=date_to, customer_id=current_user.id)
    else:
        raise ObjectNotFoundHTTPException
    book = Workbook()
    sheet = book.active

    headers = ["ИМЯ", "EMAIL", "РОЛЬ", "РЕЙТИНГ", "ПЕРИОД"]

    for i, header in enumerate(headers, start=1):
        cell = sheet[f"A{i}"]
        cell.value = header
        cell.font = Font(bold=True)

    sheet['B1'] = f"{current_user.first_name} {current_user.middle_name} {current_user.last_name}"
    sheet['B2'] = current_user.email
    sheet['B3'] = current_user.role
    sheet['B4'] = current_user.rating
    sheet['B4'].alignment = Alignment(horizontal="left", vertical="center")
    sheet['B5'] = f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"

    sheet.column_dimensions["B"].width = 40
    for col in ["D", "E", "F", "G", "H", "I"]:
        sheet.column_dimensions[col].width = 15
    sheet.column_dimensions["G"].width = 18
    sheet.column_dimensions["H"].width = 16

    order_headers = ["ID", "НАЗВАНИЕ", "ОПИСАНИЕ", "СТАТУС", "ЦЕНА", "ДАТА СОЗДАНИЯ"]

    for i, header in enumerate(order_headers, start=4):
        cell = sheet.cell(row=1, column=i)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = header
        cell.font = Font(bold=True)


    all_orders = len(orders)
    total_sum = 0
    completed_orders = 0
    canceled_order = 0

    row = 2
    for order in orders:
        sheet.cell(row=row, column=4).value = order.id
        sheet.cell(row=row, column=5).value = order.title
        sheet.cell(row=row, column=6).value = order.description
        sheet.cell(row=row, column=7).value = order.status
        sheet.cell(row=row, column=8).value = order.price
        sheet.cell(row=row, column=9).value = order.created_at

        price_cell = sheet.cell(row=row, column=8)
        price_cell.value = float(order.price)
        price_cell.number_format = '#,##0.00 ₽'

        date_cell = sheet.cell(row=row, column=9)
        date_cell.value = order.created_at
        date_cell.number_format = 'DD.MM.YYYY'

        if order.status == OrderStatus.cancelled:
            canceled_order += 1
        if order.status == OrderStatus.completed:
            completed_orders += 1

        total_sum += order.price
        row += 1

    row += 2
    sheet.cell(row=row, column=4).value = "ВСЕГО ЗАКАЗОВ"
    sheet.cell(row=row, column=5).value = "ЗАВЕРШЕНО"
    sheet.cell(row=row, column=6).value = "ОТМЕНЕНО"
    sheet.cell(row=row, column=7).value = "ДОЛЯ ЗАВЕРШЕННЫХ"
    sheet.cell(row=row, column=8).value = "ИТОГОВАЯ СУММА"
    sheet.cell(row=row, column=9).value = "СРЕДНИЙ ЧЕК"

    for i in range(4, 10):
        cell = sheet.cell(row=row, column=i)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row += 1

    sheet.cell(row=row, column=4).value = all_orders
    sheet.cell(row=row, column=5).value = completed_orders
    sheet.cell(row=row, column=6).value = canceled_order
    sheet.cell(row=row, column=7).value = float(completed_orders/all_orders)
    sheet.cell(row=row, column=7).number_format = '0.0%'
    sheet.cell(row=row, column=8).value = total_sum
    sheet.cell(row=row, column=8).number_format = '#,##0.00 ₽'
    sheet.cell(row=row, column=9).value = float(total_sum / all_orders)
    sheet.cell(row=row, column=9).number_format = '#,##0.00 ₽'

    book.save("besmila.xlsx")
    book.close()

    return FileResponse("besmila.xlsx", filename=f"besmila.xlsx")

@router.patch("/me")
async def partially_update_profile(
    db: DBDep,
    data: UserPatch,
    current_user: CurrentUserDep,
):
    try:
        await AuthenticationService(db).partially_update_profile(current_user.id, data)
    except PermissionDeniedException:
        raise PermissionDeniedHTTPException
    return {"status": "OK"}


@router.post("/logout")
async def logout(
    db: DBDep,
    response: Response
):
    await AuthenticationService(db).logout(response)
    return {"status": "OK"}