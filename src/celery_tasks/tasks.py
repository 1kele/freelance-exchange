import asyncio
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from src.celery_tasks.celery_app import celery_instance
from src.database import async_session_maker_null_pool
from src.db_manager import DBManager
from src.exceptions import ObjectNotFoundHTTPException
from src.schemas.order import OrderStatus
from src.schemas.user import AllRoles
from fastapi_mail import MessageSchema, MessageType, NameEmail
from src.mail import fastmail
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


@celery_instance.task
def create_report_exel(
    date_from: str,
    date_to: str,
    user_id: int,
    user_role: AllRoles,
    user_email: str,
    user_first_name: str,
    user_middle_name: str,
    user_last_name: str,
    user_rating: float,
):
    asyncio.run(
        _create_exel_report_async(
            date_from,
            date_to,
            user_id,
            user_role,
            user_email,
            user_first_name,
            user_middle_name,
            user_last_name,
            user_rating,
        )
    )


async def _create_exel_report_async(
    date_from: str,
    date_to: str,
    user_id: int,
    user_role: AllRoles,
    user_email: str,
    user_first_name: str,
    user_middle_name: str,
    user_last_name: str,
    user_rating: float,
):
    date_from_str = date_from
    date_to_str = date_to

    date_from_dt = datetime.fromisoformat(date_from_str) # type: ignore
    date_to_dt = datetime.fromisoformat(date_to_str) # type: ignore
    async with DBManager(session_factory=async_session_maker_null_pool) as db:
        if user_role == AllRoles.customer:
            orders = await db.order.get_by_date_range(
                date_from=date_from, date_to=date_to, customer_id=user_id
            )
        elif user_role == AllRoles.freelancer:
            orders = await db.order.get_filter_by(
                date_from=date_from, date_to=date_to, customer_id=user_id
            )
        else:
            raise ObjectNotFoundHTTPException
        book = Workbook()
        sheet = book.active
        assert sheet is not None

        headers = ["ИМЯ", "EMAIL", "РОЛЬ", "РЕЙТИНГ", "ПЕРИОД"]

        for i, header in enumerate(headers, start=1):
            cell = sheet[f"A{i}"]
            cell.value = header
            cell.font = Font(bold=True)

        sheet["B1"] = f"{user_first_name} {user_middle_name} {user_last_name}"
        sheet["B2"] = user_email
        sheet["B3"] = user_role
        sheet["B4"] = user_rating
        sheet["B4"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["B5"] = (
            f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}" # type: ignore
        )

        sheet.column_dimensions["B"].width = 40
        for col in ["D", "E", "F", "G", "H", "I"]:
            sheet.column_dimensions[col].width = 15
        sheet.column_dimensions["G"].width = 18
        sheet.column_dimensions["H"].width = 16

        order_headers = [
            "ID",
            "НАЗВАНИЕ",
            "ОПИСАНИЕ",
            "СТАТУС",
            "ЦЕНА",
            "ДАТА СОЗДАНИЯ",
        ]

        for i, header in enumerate(order_headers, start=4):
            cell = sheet.cell(row=1, column=i)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = header # type: ignore
            cell.font = Font(bold=True)

        all_orders = len(orders)
        total_sum = 0
        completed_orders = 0
        canceled_order = 0

        row = 2
        for order in orders:
            sheet.cell(row=row, column=4).value = order.id
            sheet.cell(row=row, column=5).value = order.title
            sheet.cell(row=row, column=6).value = order.description
            sheet.cell(row=row, column=7).value = order.status
            sheet.cell(row=row, column=8).value = order.price
            sheet.cell(row=row, column=9).value = order.created_at

            price_cell = sheet.cell(row=row, column=8)
            price_cell.value = float(order.price) # type: ignore
            price_cell.number_format = "#,##0.00 ₽"

            date_cell = sheet.cell(row=row, column=9)
            date_cell.value = order.created_at
            date_cell.number_format = "DD.MM.YYYY"

            if order.status == OrderStatus.cancelled:
                canceled_order += 1
            if order.status == OrderStatus.completed:
                completed_orders += 1

            total_sum += order.price
            row += 1

        row += 2
        sheet.cell(row=row, column=4).value = "ВСЕГО ЗАКАЗОВ" # type: ignore
        sheet.cell(row=row, column=5).value = "ЗАВЕРШЕНО" # type: ignore
        sheet.cell(row=row, column=6).value = "ОТМЕНЕНО" # type: ignore
        sheet.cell(row=row, column=7).value = "ДОЛЯ ЗАВЕРШЕННЫХ" # type: ignore
        sheet.cell(row=row, column=8).value = "ИТОГОВАЯ СУММА" # type: ignore
        sheet.cell(row=row, column=9).value = "СРЕДНИЙ ЧЕК" # type: ignore

        for i in range(4, 10):
            cell = sheet.cell(row=row, column=i)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        sheet.cell(row=row, column=4).value = all_orders # type: ignore
        sheet.cell(row=row, column=5).value = completed_orders # type: ignore
        sheet.cell(row=row, column=6).value = canceled_order # type: ignore
        sheet.cell(row=row, column=7).value = float(completed_orders / all_orders) # type: ignore
        sheet.cell(row=row, column=7).number_format = "0.0%" # type: ignore
        sheet.cell(row=row, column=8).value = total_sum # type: ignore
        sheet.cell(row=row, column=8).number_format = "#,##0.00 ₽" # type: ignore
        sheet.cell(row=row, column=9).value = float(total_sum / all_orders) # type: ignore
        sheet.cell(row=row, column=9).number_format = "#,##0.00 ₽" # type: ignore

        filename = f"report_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        book.save(filename)
        book.close()

        message = MessageSchema(
            subject="Ваш Отчет Готов",
            recipients=[NameEmail(name=user_first_name, email=user_email)],
            body=f"""
                    <html>
                    <body>
                        <ins><h2>Здравствуйте, {user_first_name}!</h2></ins>
                        <u><p>Ваш отчёт за период <b>{date_from_dt.strftime("%d.%m.%Y")}</b> — <b>{date_to_dt.strftime("%d.%m.%Y")}</b> готов.</p></u>
                        Отчет пушка он уже доступен в 
                        <br>
                        <p>Файл во вложении.</p>
                    </body>
                    </html>
                    """,
            subtype=MessageType.html,
            attachments=[filename],
        )
        try:
            await fastmail.send_message(message)
            print("EMAIL SENT")

        except Exception as e:
            print("EMAIL ERROR:", e)


async def _create_pdf_report_async(
    user_id: int,
    date_from: str,
    date_to: str,
):
    date_from_dt = datetime.fromisoformat(date_from) # type: ignore
    date_to_dt = datetime.fromisoformat(date_to) # type: ignore

    async with DBManager(session_factory=async_session_maker_null_pool) as db:
        current_user = await db.user.get_one(id=user_id)
        if current_user.role == AllRoles.customer:
            orders = await db.order.get_by_date_range(
                date_from=date_from_dt, date_to=date_to_dt, customer_id=user_id
            )

        elif current_user.role == AllRoles.freelancer:
            orders = await db.order.get_filter_by(
                date_from=date_from_dt, date_to=date_to_dt, customer_id=user_id
            )

        else:
            raise ObjectNotFoundHTTPException

        filename = f"report_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        doc = SimpleDocTemplate(
            filename,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()

        elements = []

        title = Paragraph(f"<b>User Report: {current_user.first_name}</b>", styles["Title"])

        user_info = Paragraph(
            f"""
            <b>Name:</b> {current_user.first_name} {current_user.middle_name} {current_user.last_name}<br/>
            <b>Email:</b> {current_user.email}<br/>
            <b>Role:</b> {current_user.role}<br/>
            <b>Rating:</b> {current_user.rating}<br/>
            <b>Period:</b>
            {date_from_dt.strftime("%d.%m.%Y")} —
            {date_to_dt.strftime("%d.%m.%Y")}
            """,
            styles["BodyText"],
        )

        elements.append(title)
        elements.append(Spacer(1, 20))
        elements.append(user_info)
        elements.append(Spacer(1, 20))

        data = [["ID", "Title", "Description", "Status", "Price", "Created"]]

        for order in orders:
            description = (
                order.description[:50] + "..."
                if order.description and len(order.description) > 50
                else order.description
            )

            data.append(
                [
                    str(order.id),
                    str(order.title),
                    str(description),
                    str(order.status),
                    f"{float(order.price):,.2f} RUB",
                    order.created_at.strftime("%d.%m.%Y"),
                ]
            )

        table = Table(data, repeatRows=1, colWidths=[40, 100, 180, 80, 80, 80])

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        elements.append(table)

        doc.build(elements)

        message = MessageSchema(
            subject="Your PDF Report Is Ready",
            recipients=[NameEmail(name=current_user.first_name, email=current_user.email)],
            body=f"""
            <html>
            <body>
                <h2>Hello, {current_user.first_name}!</h2>
                <p>
                    Your PDF report for the period
                    <b>{date_from_dt.strftime("%d.%m.%Y")}</b> —
                    <b>{date_to_dt.strftime("%d.%m.%Y")}</b>
                    is ready.
                </p>
                <p>The file is attached.</p>
            </body>
            </html>
            """,
            subtype=MessageType.html,
            attachments=[filename],
        )

        try:
            await fastmail.send_message(message)
            print("EMAIL SENT")

        except Exception as e:
            print("EMAIL ERROR:", e)
